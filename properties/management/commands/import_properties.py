"""
Django management command: import_properties

USAGE
------
    python manage.py import_properties path/to/dubai_projects_import_v2.xlsx
    python manage.py import_properties path/to/dubai_projects_import_v2.xlsx --dry-run

WHERE TO PUT THIS FILE
-----------------------
properties/management/commands/import_properties.py

Django needs these empty files to exist too (create them if they don't):
    properties/management/__init__.py
    properties/management/commands/__init__.py

WHAT THIS DOES
--------------
- Reads one row per project from the Excel sheet.
- For each row, in its own atomic transaction (so one bad row never corrupts
  the others):
    * looks up City by exact name (all rows use "Dubai", which already
      exists in your DB)
    * looks up DeveloperCompany by exact name (all 10 developers in this
      sheet already exist in your DB - this will link to them, not
      duplicate them)
    * looks up District by exact name+city. If district_status was NEW,
      this creates a new District row. If the district cell still reads
      "NEEDS_DECISION" (the 3 AMBIGUOUS rows), the row is SKIPPED and
      reported at the end - it will NOT create a garbage district.
    * get_or_create's each PropertyFacility named in the amenities column
    * creates (or updates, if a Property with this title already exists)
      the Property row
- Does NOT touch images. `cover` is left blank and no PropertyImage rows
  are created - add real photos through the admin afterwards.
- Does NOT create Apartment / GroupedApartment / Unit rows - the sheet's
  unit_mix column is a free-text summary, not structured per-type pricing.

EXPECTED EXCEL COLUMNS (exact header names, row 1)
---------------------------------------------------
title, developer_company, city, district, district_status, description,
property_type, price_from_aed, area_range_sqft, unit_mix, delivery_date,
sales_status, facilities_amenities, official_source_url, notes
"""

import re
import traceback
from datetime import date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from properties.models import (
    City,
    District,
    DeveloperCompany,
    PropertyStatus,
    SalesStatus,
    PropertyType,
    PropertyFacility,
    Property,
)

REQUIRED_HEADERS = [
    "title",
    "developer_company",
    "city",
    "district",
    "district_status",
    "description",
    "property_type",
    "price_from_aed",
    "area_range_sqft",
    "unit_mix",
    "delivery_date",
    "sales_status",
    "facilities_amenities",
    "official_source_url",
    "notes",
]

QUARTER_END = {"1": (3, 31), "2": (6, 30), "3": (9, 30), "4": (12, 31)}
FALLBACK_DELIVERY_DATE = date(2030, 1, 1)
NEEDS_DECISION_MARKER = "needs_decision"


def unique_slug(model, base_text, instance_pk=None):
    base = slugify(base_text)[:240] or "item"
    slug = base
    counter = 2
    qs = model.objects.all()
    while qs.filter(slug=slug).exclude(pk=instance_pk).exists():
        slug = f"{base}-{counter}"
        counter += 1
    return slug


def parse_price(raw):
    """
    '2,550,000'              -> Decimal('2550000.00')
    '2,300,000 - 2,590,000'  -> Decimal('2300000.00')  (only the first/starting number)
    'Not published - verify' -> None

    Returns a Decimal already quantized to 2 places, or None. Never returns
    a float and never lets decimal.InvalidOperation escape - Django's own
    conversion during save() is where that error was happening before, so
    doing the conversion safely here up front avoids it entirely.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None

    # A range like "1,200,000 - 1,400,000" - only take the first number,
    # never concatenate both sides together.
    first_part = re.split(r"[-\u2013\u2014]", text)[0]

    digits = re.sub(r"[^\d.]", "", first_part)
    if not digits or digits == ".":
        return None

    try:
        value = Decimal(digits)
    except InvalidOperation:
        return None

    # Property.price is DecimalField(max_digits=12, decimal_places=2) ->
    # at most 10 digits before the decimal point. Never let a bad parse
    # reach save() - just drop it and flag it instead.
    if value >= Decimal("10000000000"):
        return None

    try:
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def parse_delivery_date(raw, warnings, title):
    if raw is None:
        warnings.append(f"[{title}] no delivery_date given -> used {FALLBACK_DELIVERY_DATE}")
        return FALLBACK_DELIVERY_DATE

    text = str(raw).strip()

    m = re.search(r"Q([1-4])\s*(\d{4})", text, re.IGNORECASE)
    if m:
        month, day = QUARTER_END[m.group(1)]
        return date(int(m.group(2)), month, day)

    m = re.search(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})",
        text,
        re.IGNORECASE,
    )
    if m:
        month_names = [
            "january", "february", "march", "april", "may", "june",
            "july", "august", "september", "october", "november", "december",
        ]
        month = month_names.index(m.group(1).lower()) + 1
        return date(int(m.group(2)), month, 28)

    m = re.search(r"(\d{4})", text)
    if m:
        return date(int(m.group(1)), 12, 31)

    warnings.append(f"[{title}] could not parse delivery_date '{text}' -> used {FALLBACK_DELIVERY_DATE}")
    return FALLBACK_DELIVERY_DATE


def derive_property_status_name(sales_status_text):
    text = (sales_status_text or "").lower()
    if "ready" in text or "delivered" in text or "completed" in text:
        return "Ready"
    return "Off-Plan"


class Command(BaseCommand):
    help = "Bulk-import properties from the researched Dubai projects Excel sheet."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to the .xlsx file")
        parser.add_argument("--dry-run", action="store_true",
                             help="Parse and validate everything, but don't write to the database.")

    def handle(self, *args, **options):
        path = options["excel_path"]
        dry_run = options["dry_run"]

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {path}")

        ws = wb["Projects"] if "Projects" in wb.sheetnames else wb.active

        header_row = [c.value for c in ws[1]]
        missing = [h for h in REQUIRED_HEADERS if h not in header_row]
        if missing:
            raise CommandError(f"Excel is missing expected column(s): {missing}")

        col_index = {name: header_row.index(name) for name in REQUIRED_HEADERS}

        created, updated, skipped, warnings = [], [], [], []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row is None or all(v is None for v in row):
                continue

            def get(col):
                return row[col_index[col]]

            title = (get("title") or "").strip() if get("title") else None
            if not title:
                skipped.append(f"Row {row_num}: no title, skipped")
                continue

            district_value = (get("district") or "").strip()
            if district_value.lower().replace(" ", "_") == NEEDS_DECISION_MARKER or \
               district_value.upper() == "NEEDS_DECISION":
                skipped.append(
                    f"Row {row_num} ({title}): district still says NEEDS_DECISION - "
                    f"resolve it in the sheet and re-run this row"
                )
                continue

            try:
                with transaction.atomic():
                    self._import_row(get, title, dry_run, created, updated, warnings)
            except Exception as exc:  # noqa: BLE001
                tb = traceback.format_exc()
                skipped.append(f"Row {row_num} ({title}): FAILED - {exc}\n{tb}")

        self.stdout.write(self.style.SUCCESS(f"\nCreated: {len(created)}"))
        for line in created:
            self.stdout.write(f"  + {line}")

        self.stdout.write(self.style.WARNING(f"\nUpdated (already existed): {len(updated)}"))
        for line in updated:
            self.stdout.write(f"  ~ {line}")

        if skipped:
            self.stdout.write(self.style.ERROR(f"\nSkipped / failed: {len(skipped)}"))
            for line in skipped:
                self.stdout.write(f"  ! {line}")

        if warnings:
            self.stdout.write(self.style.WARNING(f"\nNeeds a manual look ({len(warnings)}):"))
            for line in warnings:
                self.stdout.write(f"  ? {line}")

        self.stdout.write(self.style.WARNING(
            "\nReminder: no images were imported. Add cover photos and gallery "
            "images for each property through the admin."
        ))
        if dry_run:
            self.stdout.write(self.style.NOTICE("\nDRY RUN - nothing was written to the database."))

    def _import_row(self, get, title, dry_run, created, updated, warnings):
        city_name = (get("city") or "Dubai").strip()
        district_name = (get("district") or "").strip()
        developer_name = (get("developer_company") or "").strip()
        property_type_name = (get("property_type") or "Apartment").strip()
        sales_status_text = (get("sales_status") or "Off-Plan").strip()
        description = (get("description") or "").strip()
        facilities_text = get("facilities_amenities") or ""
        price = parse_price(get("price_from_aed"))
        delivery_date = parse_delivery_date(get("delivery_date"), warnings, title)
        property_status_name = derive_property_status_name(sales_status_text)

        if dry_run:
            created.append(f"{title} (dry-run, not written)")
            return

        # City / Developer are expected to already exist (exact match) -
        # get_or_create still works safely either way.
        city, _ = City.objects.get_or_create(
            name=city_name, defaults={"slug": unique_slug(City, city_name)}
        )

        developer, dev_created = DeveloperCompany.objects.get_or_create(
            name=developer_name,
            defaults={"slug": unique_slug(DeveloperCompany, developer_name)},
        )
        if dev_created:
            warnings.append(f"[{title}] developer '{developer_name}' did not exist and was created new - expected an exact match, please verify")

        district, dist_created = District.objects.get_or_create(
            city=city,
            name=district_name,
            defaults={"slug": unique_slug(District, f"{district_name}-{city_name}")},
        )
        if dist_created:
            warnings.append(f"[{title}] district '{district_name}' was newly created under {city_name}")

        property_type, _ = PropertyType.objects.get_or_create(
            name=property_type_name,
            defaults={"slug": unique_slug(PropertyType, property_type_name)},
        )

        property_status, _ = PropertyStatus.objects.get_or_create(
            name=property_status_name,
            defaults={"slug": unique_slug(PropertyStatus, property_status_name)},
        )

        sales_status, _ = SalesStatus.objects.get_or_create(
            name=sales_status_text,
            defaults={
                "value": slugify(sales_status_text),
                "class_name": "badge-default",
                "mobile_color": "#808080",
            },
        )

        facility_objs = []
        for name in [f.strip() for f in str(facilities_text).split(",") if f.strip()]:
            facility, _ = PropertyFacility.objects.get_or_create(
                name=name, defaults={"slug": unique_slug(PropertyFacility, name)}
            )
            facility_objs.append(facility)

        existing = Property.objects.filter(title=title).first()

        common_fields = dict(
            description=description or "Description pending.",
            delivery_date=delivery_date,
            property_status=property_status,
            sales_status=sales_status,
            developer_company=developer,
            city=city,
            district=district,
            property_type=property_type,
            price=price,
        )

        if existing:
            for field, value in common_fields.items():
                setattr(existing, field, value)
            existing.save()
            existing.facilities.set(facility_objs)
            updated.append(title)
        else:
            prop = Property(title=title, **common_fields)
            prop.save()
            prop.facilities.set(facility_objs)
            created.append(title)

        if price is None:
            warnings.append(f"[{title}] price could not be parsed - left blank, fill in manually")